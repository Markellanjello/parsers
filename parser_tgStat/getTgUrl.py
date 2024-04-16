import openpyxl
from openpyxl.writer.excel import save_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from tgPars import get_data
FILE_NAME ="telegram.xlsx"
columns = ["Название", "Ссылка", "Описание", "Тематика", "Охват"]

try:
  wb = openpyxl.load_workbook(FILE_NAME)
except:
  wb = openpyxl.Workbook()

ws = wb.create_sheet(input("Region :"))
col = ws.column_dimensions['A'].width = 20
col1 = ws.column_dimensions['B'].width = 40
col2 = ws.column_dimensions['C'].width = 40
col3 = ws.column_dimensions['D'].width = 20
col4 = ws.column_dimensions['E'].width = 20

for i, value in enumerate(columns, 1):
  ws.cell(row=1, column=i).value = value

for i, row in enumerate(get_data(), 2):
  for j, value in enumerate(row, 1):
    ws.cell(row=i, column=j).value = value

save_workbook(wb, FILE_NAME)