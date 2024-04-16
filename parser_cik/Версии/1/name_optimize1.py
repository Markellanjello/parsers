import openpyxl
from openpyxl.writer.excel import save_workbook
from parser_name_copy import get_data
import datetime
current_date = datetime.date.today().isoformat()
reg = input("Region :")
FILE_NAME ="mane.xlsx"
columns = ["№ п/п", "ФИО кандидата", "Дата рождения кандидата", "Субъект выдвижения", "Номер округа", "Выдвижение",	"Регистрация", "Избрание"]

try:
	wb = openpyxl.load_workbook(FILE_NAME)
except:
	wb = openpyxl.Workbook()

ws = wb.create_sheet(reg + " " + current_date + " ")

for i, value in enumerate(columns, 1):
	ws.cell(row=1, column=i).value = value

for i, row in enumerate(get_data(), 2):
	for j, value in enumerate(row, 1):
		ws.cell(row=i, column=j).value = value

save_workbook(wb, FILE_NAME)



