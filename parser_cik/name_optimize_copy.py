import openpyxl
from openpyxl.writer.excel import save_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from parser_name_copy import get_data, get_head
import datetime
current_date = datetime.date.today().isoformat()
reg = input("Region :")
FILE_NAME ="mane.xlsx"
for wow in get_head():
	columns = wow
print(columns)
for wiw in get_data():
	print(wiw)

# try:
# 	wb = openpyxl.load_workbook(FILE_NAME)
# except:
# 	wb = openpyxl.Workbook()

# ws = wb.create_sheet(reg + " " + current_date + " ")
# col = ws.column_dimensions['A'].width = 8
# col1 = ws.column_dimensions['B'].width = 40
# col2 = ws.column_dimensions['C'].width = 30
# col3 = ws.column_dimensions['D'].width = 130
# col4 = ws.column_dimensions['E'].width = 30
# col5 = ws.column_dimensions['F'].width = 30
# col6 = ws.column_dimensions['G'].width = 30
# col7 = ws.column_dimensions['H'].width = 30
# col8 = ws.column_dimensions['I'].width = 30
# col9 = ws.column_dimensions['J'].width = 30
# col10 = ws.column_dimensions['K'].width = 30

# for i, value in enumerate(columns, 1):
# 	ws.cell(row=1, column=i).value = value

# for i, row in enumerate(get_data(), 2):
# 	for j, value in enumerate(row, 1):
# 		ws.cell(row=i, column=j).value = value

# save_workbook(wb, FILE_NAME)



